'''try:
    num=int(input("Enter a number:"))
    result=10/num
    print("Results:",result)
except ZeroDivisionError:#if zerodivisionerror is come then it prints the 
    #down print statement
    print("Error: Cannot divide by zero.")
except ValueError:
    #if i entered the char,float,str different dataset from int it prints the down print statement
    print("Error: Invalid input, please enter a number.")
finally:
    print("Execution finished")'''
    #the above finally block is always prints at end
'''
Enter a number:0
Error: Cannot divide by zero.
Execution finished
'''
#another example
'''
Enter a number:3.5
Error: Invalid input, please enter a number.
Execution finished
'''

#Raising the exceptions 
def check_name(name):
    if not isinstance(name,str):
        raise ValueError("Name must be in string format!")
    return name
print(check_name(name="shiva"))

#print(check_name(name=12))

def checking_name(name):
    if not name.isalpha():
        raise ValueError("Name must contain only alphabets.")
    return name
#print(checking_name("shiva123"))#it returns ValueError
print(checking_name("ganga"))
'''
Write a Python function divide_numbers(a, b) that:

Takes two inputs a and b.
Returns the result of a / b.
Raises a ValueError if either input is not a number.
Handles division by zero gracefully with a custom error message.
Always prints "Operation complete" at the end (using finally).'''
try:
    def divide_numbers(a,b):
        return a/b
    print(divide_numbers(a,10))
except ZeroDivisionError:
    print("Error:Zero Division Error ")
except ValueError:
    print("Error:Values must be numbers  ")
except NameError:
    print("Error: name is not defined")
finally:
    print("Operation completed!")

'''
Error: name is not defined
Operation completed!
'''
#using another way to implement
def divide_number(a,b):
    try:
        if not isinstance(a,(int,float)) or  not isinstance(b,(int,float)):
            #here im using the or and checking is that is num or what 
            raise ValueError("Inputs must be numbers.")
        result=a/b
        return result
    except ZeroDivisionError:
        return "Error: Cannot divide by zero."
    except ValueError as e:
        return "Error:",e
    finally:
        print("Operation complete.")
print(divide_number(10,2))
#5.0
print(divide_number(10,0))
'''
Error: Cannot divide by zero.
Operation complete.
'''
print(divide_number("tt",2))
'''
Error: Inputs must be numbers.
Operation complete.
'''
#Question-2
'''
Write a Python function read_file(filename) that:

Opens a file in read mode.
Raises a FileNotFoundError if the file does not exist.
Prints the file content if it exists.
Handles any other unexpected errors gracefully.
Always prints "File operation complete" at the end (using finally).'''


'''from openpyxl import load_workbook
def file_excel(filename):
    try:
        #trying to open the workbook
        workbook=load_workbook(filename)
        sheet=workbook.active
        #printing the first 5 rows as a sample
        for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            print(row)

    except FileNotFoundError:
        print("Error: File not found")
    except Exception as e:
        print("Unexpected error:",e)
    finally:
        print("Excel File operation complete.")
file_excel("C:/Users/saisw/OneDrive/Documents/Book1.xlsx")

'''
#it prints the first five lines
'''
Write a Python function validate_excel(filename) that:

Opens the given Excel file using openpyxl.
Reads the first column of the active sheet.
Raises a ValueError if any cell in the first column is empty.
Prints "Validation complete" at the end (using finally).
'''
'''
from openpyxl import load_workbook
def validate_excel(filename):
    try:
        workbook=load_workbook(filename)
        sheet=workbook.active
        for col in sheet.iter_cols(min_col=1, max_col=2,values_only=True):
            print(col)
    except ValueError:
        print("Error: The First column is empty.")
    finally:
        print("Validation is complete")
validate_excel("Book1.xlsx")'''
#now for validate and sum of all ages if valid
#and returns "summation complete" at end
'''
from openpyxl import load_workbook
def validate_and_sum(filename):
    try:
        workbook=load_workbook(filename)
        sheet=workbook.active
        for col in sheet.iter_cols(min_col=2,max_col=2,values_only=True):
            print(f"sum of all ages is :{sum(col)}")
    except ValueError:
        print("Error: The second col is empty.")
    finally:
        print("validation is completed")
validate_and_sum("Book1.xlsx")'''#sum of all ages is :277

#now im checking the both columns (names)-must not be empty
#and ages column must be numeric(int or float)
#raise a valueerror if the any one fails
#if all rows are valid prints the rows in format
#and at last return "data validation complete" in finally block
'''
from openpyxl import load_workbook
def validate_excel_data(filename):
    try:
        workbook=load_workbook(filename)
        sheet=workbook.active
        for row in sheet.iter_rows(min_col=1,max_col=2,values_only=True):
            name,age=row
            if name is  None:
                raise ValueError("Empty cell found in first column")
            if not isinstance(age, (int, float)):
                raise ValueError("Age must be a number")
            print(f"Name: {name}, Age: {age}")
    except FileNotFoundError:
        print("Error: File not found.")
    except ValueError as e:
        print(f"Error: {e}")
    except Exception as e:
        print("Unexpected error:", e)
    finally:
        print("Data Validation Complete.")
validate_excel_data("Book1.xlsx")
'''
#for three cols

from openpyxl import load_workbook

def validate_excel_data(filename):
    errors = []
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
        
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
            name, age, city = row
            
            if name is None:
                errors.append("Empty cell found in first column")
                continue
            if city is None:
                errors.append("Empty cell found in third column")
                continue
            if not isinstance(age, (int, float)):
                errors.append("Age must be a number")
                continue
            
            print(f"Name: {name}, Age: {age}, City: {city}")
    
    except FileNotFoundError:
        print("Error: File not found.")
    except Exception as e:
        print("Unexpected error:", e)
    finally:
        if errors:
            print("Errors found:")
            for err in errors:
                print(f"- {err}")
        
        print("Data Validation Complete.")
# Run
validate_excel_data("Book1.xlsx")

'''
Modify the validator so that it skips invalid rows instead of stooping or collecting them. In other words
-> If a row fails validation(empty None , non- numeric Age, empty city),ignore it
->only print ignore it.
->At the end , show sa summary count of how many rows were valid and how many were skipped.
'''
from openpyxl import load_workbook

def validate_excel_data_skip_invalid(filename):
    valid_count = 0
    invalid_count = 0
    
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
        
        # Skip header row 
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True), start=2):
            name, age, city = row
            
            # Validation checks and counts the invalid ones
            if name is None or city is None or not isinstance(age, (int, float)):
                invalid_count += 1
                continue  # ✅ Skip invalid row
            
            print(f"Name: {name}, Age: {age}, City: {city}")
            valid_count += 1
    
    except FileNotFoundError:
        print("Error: File not found.")
    except Exception as e:
        print("Unexpected error:", e)
    finally:
        print(f"\nSummary: Valid rows = {valid_count}, Skipped rows = {invalid_count}")
        print("Data Validation Complete.")
validate_excel_data_skip_invalid("Book1.xlsx")
'''
Name: Shiva, Age: 25, City: HYD
Name: Ravi, Age: 30, City: WGL
Name: Kumar, Age: 22, City: HNK
Name: Meena, Age: 27, City: HNK
Name: Arjun, Age: 29, City: WGL
Name: Sneha, Age: 24, City: SEC
Name: Kavya, Age: 26, City: KZJ

Summary: Valid rows = 7, Skipped rows = 3
Data Validation Complete.'''

'''
keep the current validation rules (Name not empty,Age numeric ,City not empty)
Create a new sheet called "ValidRows" and another called "InvalidRows"
Write all valid  rows into "ValidRows" with headers(Name, Age, City)
Write all invalid rows into "InvalidRows" with headers (Name, Age, City)
At the end, print a summary of how many rows were valid and invalid.
'''

from openpyxl import load_workbook, Workbook

def validate_and_split(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active
    # Create new sheets
    if "ValidRows" in workbook.sheetnames:
        valid_ws = workbook["ValidRows"]
    else:
        valid_ws = workbook.create_sheet("ValidRows")
    if "InvalidRows" in workbook.sheetnames:
        invalid_ws = workbook["InvalidRows"]
    else:
        invalid_ws = workbook.create_sheet("InvalidRows")
    # Add headers
    valid_ws.append(("Name", "Age", "City"))
    invalid_ws.append(("Row", "Name", "Age", "City", "Error"))
    valid_count = 0
    invalid_count = 0
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True), start=2):
        name, age, city = row
        if name is None:
            invalid_ws.append((row_idx, name, age, city, "Empty cell found in first column"))
            invalid_count += 1
            continue
        if city is None:
            invalid_ws.append((row_idx, name, age, city, "Empty cell found in third column"))
            invalid_count += 1
            continue
        if not isinstance(age, (int, float)):
            invalid_ws.append((row_idx, name, age, city, "Age must be a number"))
            invalid_count += 1
            continue
        
        valid_ws.append((name, age, city))
        valid_count += 1
    workbook.save("UPGRADED_Book1.xlsx")
    print(f"Valid rows: {valid_count}, Invalid rows: {invalid_count}")
    print("Validation complete. Results written to 'ValidRows' and 'InvalidRows' sheets.")
validate_and_split("Book1.xlsx")











